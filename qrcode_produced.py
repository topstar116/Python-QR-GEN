# modules needed
import qrcode
from tkinter import filedialog
from tkinter import *
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook

#select the excel file to be read
# the texts must be in the "A" column starting with "2" row. In the B column, qrcodes will be seen.
print('select xlsx, xlsm file:')
root = Tk()
root.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("xlsm files","*.xlsm"),("xlsx files","*.xlsx"),("all files","*.*")))
print (root.filename)

# select the folder to save qrcodes as png format images and excel file with qrcodes
# print('where to save excel file and qrcodes:')
# root2 = Tk()
# root2.withdraw()
# folder_selected = filedialog.askdirectory()
folder_selected = "./output"

# read the excel file
workbook = load_workbook(str(root.filename))
# sheet = workbook.active

for sheet in workbook.worksheets:

  
  # choose just shop sheet
  if sheet.title == "Sheet1" or sheet.title == "第二" or sheet.title == "緊急" or sheet.title == "合計" or sheet.title == "振り分け設定" or sheet.title == "文字列リスト(削除禁止)":
    continue

  print(sheet.title)

  # settings for qrcode to be produced
  qr = qrcode.QRCode(
      version=1,
      error_correction=qrcode.constants.ERROR_CORRECT_L,
      box_size=4,
      border=2,)

  # check QRCode Cell
  cell_range = sheet['A3':'M3']
  row_range = ['','A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
  U1 = 'H'
  user1_n = 8
  U2 = 'I'
  user2_n = 9
  Font = 'J'
  Size = 'K'
  QR1 = 'L'
  QR2 = 'M'
  

  for each_column in cell_range:
    for cell in each_column:
        if cell.value == "1点目":
          print(row_range[cell.column])
          U1 = row_range[cell.column]
          user1_n = cell.column
        if cell.value == "2点目":
          print(row_range[cell.column])
          U2 = row_range[cell.column]
          user2_n = cell.column
        if cell.value == "フォント":
          print(row_range[cell.column])
          Font = row_range[cell.column]
        if cell.value == "BOXプリント" or cell.value == "文字サイズ":
          print(row_range[cell.column])
          Size = row_range[cell.column]
        if cell.value == "1点目QR":
          print(row_range[cell.column])
          QR1 = row_range[cell.column]
        if cell.value == "2点目QR":
          print(row_range[cell.column])
          QR2 = row_range[cell.column]



  # excel file cell size settings that will be produced 
  sheet.column_dimensions[QR1].width = 25
  sheet.column_dimensions[QR2].width = 25

  for i in range(4,len(sheet['A'])+1):
    sheet.row_dimensions[i].height=150

  # Title of B column
  # sheet["B1"]="Qr_Codes"

  # production of qrcodes for each row in the A column except first row. Skips the empty rows.
  for i in range(4,len(sheet['A'])+1):
    qr.clear()
    qrcode_str = ''
    for j in range(2,user1_n):
      if sheet.cell(row=i, column=j).value is None:
        continue
      elif j == 2 or j == 4 or j == 5 or j == 7:
        continue 
      else:
        qrcode_str += str(sheet.cell(row=i, column=j).value).strip()
        qrcode_str += "&"

    # USER1
    qrcode_str += str(sheet.cell(row=i, column=user1_n).value)
    qr.add_data(qrcode_str, optimize=0)
    qr.make(fit=True)
    img = qr.make_image()
    img.save(folder_selected + "/" + str(sheet.title) +"1_"+str(i)+"_qrcode.png")
    img=openpyxl.drawing.image.Image(folder_selected + "/" + str(sheet.title) +"1_"+str(i)+"_qrcode.png")
    img.anchor = QR1 + str(i)
    sheet.add_image(img)
    qr.clear()

    # USER2
    qrcode_str += str(sheet.cell(row=i, column=user2_n).value)
    qr.add_data(qrcode_str, optimize=0)
    qr.make(fit=True)
    img = qr.make_image()
    img.save(folder_selected + "/" + str(sheet.title) +"2_"+str(i)+"_qrcode.png")
    img=openpyxl.drawing.image.Image(folder_selected + "/" + str(sheet.title) +"2_"+str(i)+"_qrcode.png")
    img.anchor = QR2 + str(i)
    sheet.add_image(img)
    qr.clear()






    sheet[QR1 + str(i)].alignment = Alignment(horizontal='center', vertical='center')
    sheet[QR2 + str(i)].alignment = Alignment(horizontal='center', vertical='center')


  # saving the excel file
  workbook.save(folder_selected+ "/qrcode_produced.xlsx")